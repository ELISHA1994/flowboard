"""
Service layer for analytics and reporting functionality.
"""
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Any, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func, or_
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.db.models import (
    Task, TaskStatus, TaskPriority, Project, ProjectRole, Category, Tag, TimeLog
)


class AnalyticsService:
    """Service for generating analytics and reports."""
    
    @staticmethod
    def get_task_statistics(
        db: Session,
        user_id: str,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        project_id: Optional[str] = None
    ) -> Dict[str, Any]:
        """Get overall task statistics for a user."""
        # Default to last 30 days if no date range specified
        if not end_date:
            end_date = datetime.now(timezone.utc)
        if not start_date:
            start_date = end_date - timedelta(days=30)
        
        # Base query for user's tasks
        base_query = db.query(Task).filter(
            or_(
                Task.user_id == user_id,
                Task.assigned_to_id == user_id
            ),
            Task.created_at >= start_date,
            Task.created_at <= end_date
        )
        
        if project_id:
            base_query = base_query.filter(Task.project_id == project_id)
        
        # Get task counts by status
        status_counts = dict(
            db.query(Task.status, func.count(Task.id))
            .filter(
                or_(
                    Task.user_id == user_id,
                    Task.assigned_to_id == user_id
                ),
                Task.created_at >= start_date,
                Task.created_at <= end_date
            )
            .group_by(Task.status)
            .all()
        )
        
        # Get task counts by priority
        priority_counts = dict(
            db.query(Task.priority, func.count(Task.id))
            .filter(
                or_(
                    Task.user_id == user_id,
                    Task.assigned_to_id == user_id
                ),
                Task.created_at >= start_date,
                Task.created_at <= end_date
            )
            .group_by(Task.priority)
            .all()
        )
        
        # Calculate completion rate
        total_tasks = base_query.count()
        completed_tasks = base_query.filter(Task.status == TaskStatus.DONE).count()
        completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
        
        # Calculate average completion time
        completed_with_time = db.query(
            func.avg(
                func.julianday(Task.completed_at) - func.julianday(Task.created_at)
            )
        ).filter(
            or_(
                Task.user_id == user_id,
                Task.assigned_to_id == user_id
            ),
            Task.status == TaskStatus.DONE,
            Task.completed_at.isnot(None),
            Task.created_at >= start_date,
            Task.created_at <= end_date
        ).scalar()
        
        avg_completion_days = float(completed_with_time) if completed_with_time else 0
        
        # Get overdue tasks
        overdue_tasks = base_query.filter(
            Task.due_date < datetime.now(timezone.utc),
            Task.status != TaskStatus.DONE
        ).count()
        
        return {
            "total_tasks": total_tasks,
            "status_breakdown": {
                "todo": status_counts.get(TaskStatus.TODO, 0),
                "in_progress": status_counts.get(TaskStatus.IN_PROGRESS, 0),
                "done": status_counts.get(TaskStatus.DONE, 0)
            },
            "priority_breakdown": {
                "low": priority_counts.get(TaskPriority.LOW, 0),
                "medium": priority_counts.get(TaskPriority.MEDIUM, 0),
                "high": priority_counts.get(TaskPriority.HIGH, 0),
                "urgent": priority_counts.get(TaskPriority.URGENT, 0)
            },
            "completion_rate": round(completion_rate, 2),
            "average_completion_days": round(avg_completion_days, 2),
            "overdue_tasks": overdue_tasks,
            "date_range": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat()
            }
        }
    
    @staticmethod
    def get_productivity_trends(
        db: Session,
        user_id: str,
        period: str = "week",  # week, month, quarter
        lookback: int = 4  # Number of periods to look back
    ) -> Dict[str, Any]:
        """Get productivity trends over time."""
        end_date = datetime.now(timezone.utc)
        
        if period == "week":
            period_days = 7
        elif period == "month":
            period_days = 30
        elif period == "quarter":
            period_days = 90
        else:
            period_days = 7
        
        trends = []
        
        for i in range(lookback):
            period_end = end_date - timedelta(days=i * period_days)
            period_start = period_end - timedelta(days=period_days)
            
            # Tasks created in period
            created_count = db.query(func.count(Task.id)).filter(
                or_(
                    Task.user_id == user_id,
                    Task.assigned_to_id == user_id
                ),
                Task.created_at >= period_start,
                Task.created_at < period_end
            ).scalar()
            
            # Tasks completed in period
            completed_count = db.query(func.count(Task.id)).filter(
                or_(
                    Task.user_id == user_id,
                    Task.assigned_to_id == user_id
                ),
                Task.completed_at >= period_start,
                Task.completed_at < period_end
            ).scalar()
            
            # Hours logged in period
            hours_logged = db.query(func.coalesce(func.sum(TimeLog.hours), 0)).filter(
                TimeLog.user_id == user_id,
                TimeLog.logged_at >= period_start,
                TimeLog.logged_at < period_end
            ).scalar()
            
            trends.append({
                "period_start": period_start.isoformat(),
                "period_end": period_end.isoformat(),
                "tasks_created": created_count,
                "tasks_completed": completed_count,
                "hours_logged": float(hours_logged)
            })
        
        # Reverse to show oldest to newest
        trends.reverse()
        
        return {
            "period_type": period,
            "trends": trends
        }
    
    @staticmethod
    def get_time_tracking_report(
        db: Session,
        user_id: str,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None,
        group_by: str = "task"  # task, project, category, day
    ) -> Dict[str, Any]:
        """Get detailed time tracking report."""
        if not end_date:
            end_date = datetime.now(timezone.utc)
        if not start_date:
            start_date = end_date - timedelta(days=30)
        
        # Base query for time logs
        base_query = db.query(TimeLog).filter(
            TimeLog.user_id == user_id,
            TimeLog.logged_at >= start_date,
            TimeLog.logged_at <= end_date
        )
        
        # Total hours logged
        total_hours = db.query(func.coalesce(func.sum(TimeLog.hours), 0)).filter(
            TimeLog.user_id == user_id,
            TimeLog.logged_at >= start_date,
            TimeLog.logged_at <= end_date
        ).scalar()
        
        entries = []
        
        if group_by == "task":
            # Group by task
            task_times = db.query(
                Task.id,
                Task.title,
                func.sum(TimeLog.hours).label("total_hours"),
                func.count(TimeLog.id).label("log_count")
            ).join(
                TimeLog, TimeLog.task_id == Task.id
            ).filter(
                TimeLog.user_id == user_id,
                TimeLog.logged_at >= start_date,
                TimeLog.logged_at <= end_date
            ).group_by(Task.id, Task.title).all()
            
            for task_id, title, hours, count in task_times:
                entries.append({
                    "task_id": task_id,
                    "task_title": title,
                    "total_hours": float(hours),
                    "log_count": count
                })
        
        elif group_by == "project":
            # Group by project
            project_times = db.query(
                Project.id,
                Project.name,
                func.sum(TimeLog.hours).label("total_hours"),
                func.count(TimeLog.id).label("log_count")
            ).join(
                Task, Task.project_id == Project.id
            ).join(
                TimeLog, TimeLog.task_id == Task.id
            ).filter(
                TimeLog.user_id == user_id,
                TimeLog.logged_at >= start_date,
                TimeLog.logged_at <= end_date
            ).group_by(Project.id, Project.name).all()
            
            for project_id, name, hours, count in project_times:
                entries.append({
                    "project_id": project_id,
                    "project_name": name,
                    "total_hours": float(hours),
                    "log_count": count
                })
        
        elif group_by == "day":
            # Group by day
            daily_times = db.query(
                func.date(TimeLog.logged_at).label("date"),
                func.sum(TimeLog.hours).label("total_hours"),
                func.count(TimeLog.id).label("log_count")
            ).filter(
                TimeLog.user_id == user_id,
                TimeLog.logged_at >= start_date,
                TimeLog.logged_at <= end_date
            ).group_by(func.date(TimeLog.logged_at)).all()
            
            for date, hours, count in daily_times:
                entries.append({
                    "date": date,
                    "total_hours": float(hours),
                    "log_count": count
                })
        
        return {
            "total_hours": float(total_hours),
            "group_by": group_by,
            "entries": entries,
            "date_range": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat()
            }
        }
    
    @staticmethod
    def get_category_distribution(
        db: Session,
        user_id: str,
        project_id: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """Get task distribution by categories."""
        query = db.query(
            Category.id,
            Category.name,
            Category.color,
            func.count(Task.id).label("task_count")
        ).join(
            Task.categories
        ).filter(
            or_(
                Task.user_id == user_id,
                Task.assigned_to_id == user_id
            )
        )
        
        if project_id:
            query = query.filter(Task.project_id == project_id)
        
        results = query.group_by(Category.id, Category.name, Category.color).all()
        
        return [
            {
                "category_id": cat_id,
                "category_name": name,
                "color": color,
                "task_count": count
            }
            for cat_id, name, color, count in results
        ]
    
    @staticmethod
    def get_tag_distribution(
        db: Session,
        user_id: str,
        project_id: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """Get task distribution by tags."""
        query = db.query(
            Tag.id,
            Tag.name,
            Tag.color,
            func.count(Task.id).label("task_count")
        ).join(
            Task.tags
        ).filter(
            or_(
                Task.user_id == user_id,
                Task.assigned_to_id == user_id
            )
        )
        
        if project_id:
            query = query.filter(Task.project_id == project_id)
        
        results = query.group_by(Tag.id, Tag.name, Tag.color).all()
        
        return [
            {
                "tag_id": tag_id,
                "tag_name": name,
                "color": color,
                "task_count": count
            }
            for tag_id, name, color, count in results
        ]
    
    @staticmethod
    def export_tasks_csv(
        db: Session,
        user_id: str,
        task_ids: Optional[List[str]] = None
    ) -> str:
        """Export tasks to CSV format."""
        query = db.query(Task).filter(
            or_(
                Task.user_id == user_id,
                Task.assigned_to_id == user_id
            )
        )
        
        if task_ids:
            query = query.filter(Task.id.in_(task_ids))
        
        tasks = query.all()
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            "ID", "Title", "Description", "Status", "Priority",
            "Created At", "Due Date", "Completed At",
            "Estimated Hours", "Actual Hours",
            "Categories", "Tags", "Assigned To", "Project"
        ])
        
        # Write task data
        for task in tasks:
            writer.writerow([
                task.id,
                task.title,
                task.description or "",
                task.status.value,
                task.priority.value,
                task.created_at.isoformat() if task.created_at else "",
                task.due_date.isoformat() if task.due_date else "",
                task.completed_at.isoformat() if task.completed_at else "",
                task.estimated_hours or 0,
                task.actual_hours or 0,
                ", ".join([cat.name for cat in task.categories]),
                ", ".join([tag.name for tag in task.tags]),
                task.assigned_to.username if task.assigned_to else "",
                task.project.name if task.project else ""
            ])
        
        return output.getvalue()
    
    @staticmethod
    def export_tasks_excel(
        db: Session,
        user_id: str,
        task_ids: Optional[List[str]] = None
    ) -> bytes:
        """Export tasks to Excel format."""
        query = db.query(Task).filter(
            or_(
                Task.user_id == user_id,
                Task.assigned_to_id == user_id
            )
        )
        
        if task_ids:
            query = query.filter(Task.id.in_(task_ids))
        
        tasks = query.all()
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write header
        headers = [
            "ID", "Title", "Description", "Status", "Priority",
            "Created At", "Due Date", "Completed At",
            "Estimated Hours", "Actual Hours",
            "Categories", "Tags", "Assigned To", "Project"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write task data
        for row, task in enumerate(tasks, 2):
            ws.cell(row=row, column=1, value=task.id)
            ws.cell(row=row, column=2, value=task.title)
            ws.cell(row=row, column=3, value=task.description or "")
            ws.cell(row=row, column=4, value=task.status.value)
            ws.cell(row=row, column=5, value=task.priority.value)
            ws.cell(row=row, column=6, value=task.created_at.isoformat() if task.created_at else "")
            ws.cell(row=row, column=7, value=task.due_date.isoformat() if task.due_date else "")
            ws.cell(row=row, column=8, value=task.completed_at.isoformat() if task.completed_at else "")
            ws.cell(row=row, column=9, value=task.estimated_hours or 0)
            ws.cell(row=row, column=10, value=task.actual_hours or 0)
            ws.cell(row=row, column=11, value=", ".join([cat.name for cat in task.categories]))
            ws.cell(row=row, column=12, value=", ".join([tag.name for tag in task.tags]))
            ws.cell(row=row, column=13, value=task.assigned_to.username if task.assigned_to else "")
            ws.cell(row=row, column=14, value=task.project.name if task.project else "")
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    @staticmethod
    def get_team_performance(
        db: Session,
        project_id: str,
        user_id: str
    ) -> Dict[str, Any]:
        """Get team performance metrics for a project."""
        # Verify user has access to project
        project = db.query(Project).filter(Project.id == project_id).first()
        if not project or not project.has_permission(user_id, ProjectRole.VIEWER):
            return {}
        
        # Get all team members
        team_members = []
        
        # Add owner first
        owner_stats = {
            "user_id": project.owner_id,
            "username": project.owner.username,
            "role": "owner",
            "tasks_assigned": 0,
            "tasks_completed": 0,
            "hours_logged": 0
        }
        
        # Count owner's assigned tasks
        owner_stats["tasks_assigned"] = db.query(func.count(Task.id)).filter(
            Task.project_id == project_id,
            Task.assigned_to_id == project.owner_id
        ).scalar()
        
        # Count owner's completed tasks
        owner_stats["tasks_completed"] = db.query(func.count(Task.id)).filter(
            Task.project_id == project_id,
            Task.assigned_to_id == project.owner_id,
            Task.status == TaskStatus.DONE
        ).scalar()
        
        # Sum owner's hours logged
        hours = db.query(func.coalesce(func.sum(TimeLog.hours), 0)).join(
            Task, TimeLog.task_id == Task.id
        ).filter(
            Task.project_id == project_id,
            TimeLog.user_id == project.owner_id
        ).scalar()
        owner_stats["hours_logged"] = float(hours)
        
        team_members.append(owner_stats)
        
        # Add other members
        for member in project.members:
            member_stats = {
                "user_id": member.user_id,
                "username": member.user.username,
                "role": member.role.value,
                "tasks_assigned": 0,
                "tasks_completed": 0,
                "hours_logged": 0
            }
            
            # Count assigned tasks
            member_stats["tasks_assigned"] = db.query(func.count(Task.id)).filter(
                Task.project_id == project_id,
                Task.assigned_to_id == member.user_id
            ).scalar()
            
            # Count completed tasks
            member_stats["tasks_completed"] = db.query(func.count(Task.id)).filter(
                Task.project_id == project_id,
                Task.assigned_to_id == member.user_id,
                Task.status == TaskStatus.DONE
            ).scalar()
            
            # Sum hours logged
            hours = db.query(func.coalesce(func.sum(TimeLog.hours), 0)).join(
                Task, TimeLog.task_id == Task.id
            ).filter(
                Task.project_id == project_id,
                TimeLog.user_id == member.user_id
            ).scalar()
            member_stats["hours_logged"] = float(hours)
            
            team_members.append(member_stats)
        
        return {
            "project_id": project_id,
            "project_name": project.name,
            "team_members": team_members
        }